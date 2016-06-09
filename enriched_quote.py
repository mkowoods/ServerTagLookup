# import server_tag_lookup as stl
import json
import openpyxl as xl
import csv
import pprint
import time
import logging

sample_data = json.load(open('./test_data/sample_api_output.json', 'rb'))
res = sample_data[0]


# found solution @ http://stackoverflow.com/questions/8469665/saving-openpyxl-file-via-text-and-filestream
# using an at the time undocumented method...
# TODO: need to refactor into the enriched quotes module
# max_len_cols = [5, 5, 5, 5, 5]  # min character width set to 5
# wb = xl.Workbook()
# ws = wb.active
# for row in rs:
#     ws.append(row)
#     col_len_arr = map(len, row)
#     for i in range(len(col_len_arr)):
#         max_len_cols[i] = max(col_len_arr[i], max_len_cols[i])
#
# for i, col_len in enumerate(max_len_cols):
#     print xl.cell.get_column_letter(i + 1), col_len
#     ws.column_dimensions[xl.cell.get_column_letter(i + 1)].width = col_len




class QuoteWorkbook:
    def __init__(self, quote_data):
        self.quote_data = quote_data
        self.wb = xl.Workbook()
        self.ws = self.wb.active
        self.num_cols = 4
        self.min_col_width = 5
        self.starting_cell = self.ws.cell('B2')
        self.current_cell = self.starting_cell

    @property
    def file_stream(self):
        return xl.writer.excel.save_virtual_workbook(self.wb)

    @property
    def _empty_row(self):
        return [''] * (self.num_cols + int(self.empty_first_column))

    def _add_row(self, cell, iterable):
        for i, val in enumerate(iterable):
            cell.offset(row=0, column=i).value = val

    def _update_border(self, cell):
        top, left, right, bottom = cell.border, None, None, None

    def _set_column_widths(self, sniff_depth=25):
        for i in range(self.num_cols):
            cell = self.starting_cell.offset(column=i)
            max_len = self.min_col_width
            for j in range(sniff_depth):
                cell_value = cell.offset(row=j).value
                if cell_value is None:
                    pass
                else:
                    max_len = max(max_len, len(cell_value))
            print cell, max_len
            self.ws.column_dimensions[cell.column].width = max_len

    def _draw_bounding_box(self, cell, width=2, height=2):
        side = xl.styles.Side(border_style='thin')
        max_x, max_y = width - 1, height - 1
        for r in range(height):
            for c in range(width):
                local_cell = cell.offset(row=r, column=c)
                if (r, c) == (0, 0):
                    local_cell.border = xl.styles.Border(top=side, left=side)
                elif (r, c) == (max_y, 0):
                    local_cell.border = xl.styles.Border(bottom=side, left=side)
                elif (r, c) == (max_y, max_x):
                    local_cell.border = xl.styles.Border(bottom=side, right=side)
                elif (r, c) == (0, max_x):
                    local_cell.border = xl.styles.Border(top=side, right=side)
                elif r == 0:
                    local_cell.border = xl.styles.Border(top=side)
                elif r == max_y:
                    local_cell.border = xl.styles.Border(bottom=side)
                elif c == 0:
                    local_cell.border = xl.styles.Border(left=side)
                elif c == max_x:
                    local_cell.border = xl.styles.Border(right=side)

    def process_quote(self):
        """
            quote_data: list of tag_data dicts such as can be found in the results of the server_tag_api
        """
        rows = []
        for line in self.quote_data:
            self.server_table(line)
            self.current_cell = self.current_cell.offset(row=1)
            print 'processed  tag: %s' % line['tag']

        self._set_column_widths()

    def server_table(self, tag_data):
        start = time.time()

        server_start = self.current_cell
        self.current_cell.offset(row=0, column=0).value = 'Product Number'
        self.current_cell.offset(row=0, column=1).value = tag_data['product_number']
        self.current_cell.offset(row=1, column=0).value = 'Tag'
        self.current_cell.offset(row=1, column=1).value = tag_data['tag']

        # Draw Bounding Box

        self.current_cell = self.current_cell.offset(row=3)
        self._add_row(self.current_cell, ['Part Number', 'Qty', 'Category SubType', 'Description'])

        r = 0
        for i, item_dict in enumerate(tag_data['components']):
            if item_dict['is_included']:
                self.current_cell = self.current_cell.offset(row=1)
                r += 1
                self._add_row(self.current_cell, [item_dict['part_number'], item_dict['qty'],
                                                  item_dict['cat_sub_type'], item_dict['description']
                                                  ])
        print 'sever_start', server_start
        self._draw_bounding_box(server_start, width=4, height=(3 + r + 1))
        self._draw_bounding_box(server_start, width=2, height=2)
        self.current_cell = self.current_cell.offset(row=(i + 2))

        logging.debug('Time to format and write server tag %s - %.2f') % (tag_data['tag'], time.time() - start)

    def depr_server_table(self, tag_data):
        # print tag_data.keys()
        # System level
        rows = []
        rows.append(['', 'Product Number', tag_data['product_number'], '', ''])
        rows.append(['', 'Tag', tag_data['tag'], '', ''])
        rows.append(self._empty_row)

        # Component Level
        rows.append(['', 'Part Number', 'Qty', 'Category SubType', 'Description'])

        for item_dict in tag_data['components']:
            if item_dict['is_included']:
                rows.append(['',
                             item_dict['part_number'],
                             item_dict['qty'],
                             item_dict['cat_sub_type'],
                             item_dict['description']
                             ])

        return rows


def run_online_test():
    manf = 'dell'
    tags = [
        'DGDHXP1',
        #       '9DKYWQ1',
        #      '978SRR1',
        'HHXC0L1'
    ]

    servers = stl.ServerTagLookUp(tags=tags, manf=manf)
    data = servers.results
    print len(data)
    # json.dump(data, open('sample_eq_tuning.json'))
    rows = process_quote(data)
    return rows


def run_offline_test():
    pass


if __name__ == "__main__":
    wrkbk = QuoteWorkbook([res])
    wrkbk.process_quote()
    wrkbk.wb.save('tmp2.xlsx')
